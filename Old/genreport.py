import os
import re
import datetime
import sys
# from openpyxl import *
import openpyxl as xl
from C35xx.parse import *
from C35xx.getrawdata import *


##  txt 파일 Array 변환 ##

def get_switch_info(file):
    sw = list
    switch = list()
    f = open(file, 'r')
    # print(f)
    if f:
        s = f.read()
        # print(s)
    else:
        print('File %s not found' % (file))
    f.close()
    if s:
        sw = s.split('\n')
        # print(sw)
        for i in sw:
            i = ' '.join(i.split())
            # print(i)
            if i != '':
                tmp = i.split(' ')
                switch.append(tmp)
                # print(tmp)
                # print(switch)
    print(switch)
    return switch


def gensimplerep(switch, anchor):
    wb = xl.load_workbook('report.xlsx')
    ws = wb['simple_report']
    ws['D'+str(anchor)] = switch['model']
    ws['G'+str(anchor)] = switch['hostname']
    ws['L'+str(anchor)] = switch['ip']
    ws['Q'+str(anchor)] = switch['cpu util']
    ws['T'+str(anchor)] = switch['mem util']
    ws['V'+str(anchor)] = switch['temperature']
    ws['Y'+str(anchor)] = switch['power']
    ws['AC'+str(anchor)] = switch['fan']
    wb.save('report.xlsx')
    wb.close()


def repmain():
    now = 'Report Date: ' + str(datetime.datetime.now())
    wb = xl.load_workbook(filename='report.xlsx')
    ws = wb['simple_report']
    ws['A2'] = now
    wb.save('report.xlsx')
    wb.close()
    anchor = 4                           # initial position.
    sw = get_switch_info('switch.txt')
    # print(sw)
    for i in sw:
        print(anchor-3)
        protocol = i[4]
        data = C35xxGetRawData(i)
        if protocol == 'ssh':
            raw = data.get_ssh()
        elif protocol == 'telnet':
            raw = data.get_telnet()
        else:
            print('Not supported!!!\n')
        del data
        switch = C35xxParse(raw)
        model = switch.getmodel()
        p = re.compile('2950')
        m = p.search(model['model'])
        if m:
            del switch
            switch = C2950Parse(raw)
        gensimplerep(switch.saveresult(), anchor)
        anchor += 1
        del switch


if __name__ == '__main__':
    print('+-------------------------------------------------------------+')
    print('| Cisco L2 (35xx, 29xx) switch simple maintenance tool...     |')
    print('| version 0.0.1                                               |')
    print('| Scripted by snowffox, Nov. 2018                             |')
    print('+-------------------------------------------------------------+')
    repmain()
    print('Done!!!')
